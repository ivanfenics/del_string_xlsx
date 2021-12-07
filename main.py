from openpyxl import load_workbook
import os
import shutil


def get_list_of_files(input_cwd):
    # files = os.listdir(cwd)
    # xls_list = filter(lambda x: x.endswith('.xlsx') or x.endswith('.xls'), files)
    # return xls_list
    xlsx_list = []
    for root, dirs, files in os.walk(input_cwd):
        for name in files:
            if name.endswith('xlsx'):
                added_str = f'{root}\\{name}'
                xlsx_list.append(added_str.replace(input_cwd, ''))
    return xlsx_list


def check_first_sheet(sheets):
    if sheets[0] == 'Содержание':
        return True
    return False


def delete_string(l_of_files, cwd):
    for f in l_of_files:
        cur_book = load_workbook(f'{cwd}\\Input\\{f}')
        print(f'####### Обработка {f[1:]} ########')
        sheets = cur_book.sheetnames
        skip = check_first_sheet(sheets=sheets)
        for sh in cur_book:
            if skip:
                skip = False
                continue
            sh.delete_rows(1)
            print(f'Удалена строка в файле {f[1:]} на листе "{sh.title[:-1]}".')
        cur_book.save(f'{cwd}\\Output{f}')


def create_output_dir(cwd, lst):
    set_of_dir_to_create = set()
    for el in lst:
        dir_name = el.split('\\')[1:-1]
        st = '\\'.join(dir_name)
        set_of_dir_to_create.add(st)
    if os.path.exists(f'{cwd}\\Output'):
        # print(f'Directory {cwd}\\Output is delete!')
        shutil.rmtree(f'{cwd}\\Output')
        for dirs in set_of_dir_to_create:
            # print(f'#####{cwd}\\Output\\{dirs}')
            os.mkdir(f'{cwd}\\Output\\{dirs}')
    else:
        for dirs in set_of_dir_to_create:
            os.mkdir(f'{cwd}\\Output\\{dirs}')


def work_with_files():
    cwd = f'{os.getcwd()}'
    list_of_files = get_list_of_files(f'{cwd}\\Input')
    # print(*list_of_files, sep='\n')
    create_output_dir(cwd, list_of_files)
    delete_string(list_of_files, cwd)


if __name__ == '__main__':
    work_with_files()
